from openpyxl import load_workbook
from Corpus.TurkishSplitter import TurkishSplitter

wb = load_workbook('Beyazperde.com Reviews.xlsx')
ws = wb.active

print('Counting number of movies...')

moviesCount = 1
for row in ws:
  moviesCount += 1

print(f'Total movies: {moviesCount}')


print('Starting Split...')

ts = TurkishSplitter()

for movie in range(2, moviesCount):
  print('--------------------------')
  print(f'Opening {movie-1} file...')
  file = open(f'{movie-1}.txt', 'a', encoding='utf-8')  
  review = ws[f'C{movie}'].value

  print(f'Writing {movie-1} review...')
  for sentence in ts.split(review):
    file.write(str(sentence) + '\n')
  print(f'Closing {movie-1} review...')
  file.close
  print('--------------------------')
